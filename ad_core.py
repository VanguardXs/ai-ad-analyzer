"""
ad_core — reusable core for the AI Ad Creative Performance Analyzer.

Takes ad performance data + creative transcripts, computes performance metrics
deterministically (Python, not the LLM), identifies winning ads, then uses an LLM
to analyze the creative of each winner and synthesize cross-ad patterns, strategic
insights and new creative concepts to test.

Exports an Excel workbook and a Google-Doc-style Markdown report.

The tool is provider-agnostic: every model call goes through get_client() and the
retry wrapper below. To switch LLM providers, change the one line in get_client()
and the model name here. (Currently wired to Groq.)
"""

import io
import json
import random
import time

import pandas as pd

# --------------------------------------------------------------------------- #
# LLM provider configuration
# --------------------------------------------------------------------------- #
TEXT_MODEL = "openai/gpt-oss-20b"   # model for creative analysis & synthesis

MAX_RETRIES = 6
BASE_BACKOFF = 4.0


def get_client(api_key: str):
    """Create the LLM client. Swap this one line to change providers."""
    from groq import Groq
    return Groq(api_key=api_key)


# --------------------------------------------------------------------------- #
# Audience configuration
# --------------------------------------------------------------------------- #
# Default persona the creative analysis and synthesis are written for. The UI
# exposes this as an editable field so a user can point the same pipeline at a
# different buyer; every prompt below takes the audience text as a parameter
# and falls back to this constant.
TARGET_AUDIENCE = (
    "High-net-worth individuals: young, ambitious, successful, and "
    "status-driven. They don't buy because of discounts; they buy for "
    "quality, exclusivity, prestige, emotional fulfillment, and a seamless "
    "experience."
)


# --------------------------------------------------------------------------- #
# Retry helper — resilient to rate limits (HTTP 429) and transient errors
# --------------------------------------------------------------------------- #
def _is_rate_limit(error: Exception) -> bool:
    if getattr(error, "status_code", None) == 429:
        return True
    text = str(error).lower()
    return any(s in text for s in ("429", "rate limit", "rate_limit", "too many requests"))


def _retry_after_seconds(error: Exception, attempt: int) -> float:
    response = getattr(error, "response", None)
    if response is not None:
        headers = getattr(response, "headers", {}) or {}
        retry_after = headers.get("retry-after") or headers.get("Retry-After")
        if retry_after:
            try:
                return float(retry_after) + 1.0
            except ValueError:
                pass
    return BASE_BACKOFF * (2 ** (attempt - 1)) + random.uniform(0, 1.5)


def call_llm_with_retry(client, on_wait=None, **kwargs):
    last_error = None
    for attempt in range(1, MAX_RETRIES + 1):
        try:
            return client.chat.completions.create(**kwargs)
        except Exception as error:  # noqa: BLE001
            last_error = error
            if attempt == MAX_RETRIES:
                break
            wait = _retry_after_seconds(error, attempt) if _is_rate_limit(error) \
                else BASE_BACKOFF * attempt
            reason = "rate limit" if _is_rate_limit(error) else "transient error"
            (on_wait or (lambda m: print(f"    [retry] {m}")))(
                f"{reason}; retry {attempt}/{MAX_RETRIES} in {wait:.1f}s")
            time.sleep(wait)
    raise last_error


def _parse_json(raw: str) -> dict:
    raw = raw.strip()
    if raw.startswith("```"):
        raw = raw.split("```")[1]
        if raw.startswith("json"):
            raw = raw[4:]
    return json.loads(raw.strip())


# --------------------------------------------------------------------------- #
# Metrics — computed deterministically in pandas (never left to the LLM)
# --------------------------------------------------------------------------- #
REQUIRED_COLUMNS = ["ad_name", "spend", "impressions", "clicks", "conversions", "revenue"]
METRIC_CHOICES = ["ROAS", "CPA", "CTR", "CVR", "Revenue"]


def _safe_div(a, b):
    return a / b if b else 0.0


def compute_metrics(df: pd.DataFrame) -> pd.DataFrame:
    """Add CTR, CVR, CPA, ROAS, AOV columns from raw ad performance data."""
    df = df.copy()
    for col in ("spend", "impressions", "clicks", "conversions", "revenue"):
        df[col] = pd.to_numeric(df[col], errors="coerce").fillna(0)

    df["CTR"] = df.apply(lambda r: round(_safe_div(r["clicks"], r["impressions"]) * 100, 2), axis=1)
    df["CVR"] = df.apply(lambda r: round(_safe_div(r["conversions"], r["clicks"]) * 100, 2), axis=1)
    df["CPA"] = df.apply(lambda r: round(_safe_div(r["spend"], r["conversions"]), 2), axis=1)
    df["ROAS"] = df.apply(lambda r: round(_safe_div(r["revenue"], r["spend"]), 2), axis=1)
    df["AOV"] = df.apply(lambda r: round(_safe_div(r["revenue"], r["conversions"]), 2), axis=1)
    return df


def pick_winners(df: pd.DataFrame, metric: str = "ROAS", top_n: int = 3,
                 min_spend: float = 0.0) -> pd.DataFrame:
    """Rank ads and return the top performers by the chosen metric.

    For CPA lower is better; for everything else higher is better. Ads below the
    minimum spend are excluded so tiny-sample flukes don't count as winners.
    """
    pool = df[df["spend"] >= min_spend].copy()
    ascending = (metric == "CPA")
    # For CPA, ignore ads with 0 conversions (CPA would be 0 and look "best").
    if metric == "CPA":
        pool = pool[pool["conversions"] > 0]
    sort_col = "revenue" if metric == "Revenue" else metric
    ranked = pool.sort_values(sort_col, ascending=ascending)
    return ranked.head(top_n)


# --------------------------------------------------------------------------- #
# Creative analysis (per winning ad)
# --------------------------------------------------------------------------- #
# The "%%AUDIENCE%%" marker is substituted with the (possibly user-overridden)
# target audience text at call time. Plain str.replace() is used instead of
# str.format() so the literal curly braces in the JSON schema below don't need
# escaping.
AD_ANALYSIS_PROMPT = """You are an elite direct-to-consumer marketing strategist and consumer psychologist specializing in high-ticket e-commerce. Analyze the ad transcript/copy below through the lens of this TARGET AUDIENCE:

%%AUDIENCE%%

Break down WHY this ad works for that specific audience, not just what it says. Return ONLY valid JSON, no markdown, no commentary:
{
  "hook_style": {
    "angle": "the exact opening angle (e.g. status-driven, problem-solving, visual flex, aspirational)",
    "why_it_grabs_them": "why it catches an ambitious/wealthy viewer's attention"
  },
  "core_angle_mechanism": {
    "main_message": "the core message in one line",
    "must_have_positioning": "how the product is framed as a must-have / symbol of success"
  },
  "objections_handled": ["specific objections a high-standard buyer has that this ad secretly overcomes, e.g. quality verification, delivery headaches, durability, social proof"],
  "visual_script_structure": {
    "what_video_shows": "the visual pacing/what is shown",
    "what_transcript_says": "what the script is saying in parallel",
    "pacing_notes": "how the beats are structured"
  },
  "psychological_triggers": ["the prestige/status/emotional triggers used"]
}

AD TRANSCRIPT:
"""


def analyze_ad(transcript: str, client, audience: str = TARGET_AUDIENCE, on_wait=None) -> dict:
    """LLM breakdown of a single winning ad's creative, for the given audience."""
    prompt = AD_ANALYSIS_PROMPT.replace("%%AUDIENCE%%", audience) + transcript
    response = call_llm_with_retry(
        client, on_wait=on_wait, model=TEXT_MODEL, temperature=0.2,
        messages=[{"role": "user", "content": prompt}],
    )
    return _parse_json(response.choices[0].message.content)


# --------------------------------------------------------------------------- #
# Strategic synthesis (across all winners)
# --------------------------------------------------------------------------- #
SYNTHESIS_PROMPT = """You are a senior DTC growth strategist and consumer psychologist. Below are the top-performing ads for a high-ticket e-commerce brand, each with performance metrics and a strategist-grade creative breakdown. You are building strategy specifically for this TARGET AUDIENCE:

%%AUDIENCE%%

Find the PATTERNS across the winners and turn them into strategy this audience will respond to. Be specific and actionable; a creative team should be able to brief a new ad from this today. Return ONLY valid JSON, no markdown, no commentary:
{
  "winning_patterns": {
    "common_hooks": ["patterns in how winners open"],
    "common_angles": ["recurring persuasion angles"],
    "common_formats": ["recurring creative/visual formats"],
    "common_triggers": ["recurring psychological triggers"]
  },
  "why_it_works": "2-4 sentences on why these win with THIS premium audience",
  "audience_insights": {
    "status_drivers": ["what makes this audience feel and signal status"],
    "prestige_desires": ["the prestige/exclusivity outcomes this audience craves"],
    "key_objections": ["objections the brand must keep answering for a high-standard buyer"]
  },
  "new_ad_concepts": [
    {
      "concept_name": "Short name",
      "hook": "A ready-to-use opening hook aimed at the premium audience",
      "angle": "The angle it leans on",
      "format": "Suggested creative format",
      "script_outline": "3-5 short beats for the script",
      "why_it_should_work": "Tie it back to the winning patterns and premium-audience psychology",
      "what_to_test": "The single variable this concept tests"
    }
  ],
  "testing_recommendations": ["concrete next tests the team should run"]
}

Generate AT LEAST 3 items in new_ad_concepts.

DATA (winning ads, metrics, and creative breakdowns):
"""


def synthesize_insights(winners_payload: list, client, audience: str = TARGET_AUDIENCE,
                        on_wait=None) -> dict:
    """Take winners + their analyses and produce patterns, insights and new concepts."""
    payload = json.dumps(winners_payload, ensure_ascii=False, indent=2)
    prompt = SYNTHESIS_PROMPT.replace("%%AUDIENCE%%", audience) + payload
    response = call_llm_with_retry(
        client, on_wait=on_wait, model=TEXT_MODEL, temperature=0.4,
        messages=[{"role": "user", "content": prompt}],
    )
    return _parse_json(response.choices[0].message.content)


# --------------------------------------------------------------------------- #
# Excel report
# --------------------------------------------------------------------------- #
def _cell(v):
    """Coerce a value into something openpyxl can write into a single cell.

    LLM output doesn't strictly follow the requested schema, so fields typed
    as strings sometimes come back as lists or dicts; openpyxl raises on
    those, so flatten them to readable text before they ever reach a cell.
    """
    if isinstance(v, list):
        return ", ".join(str(x) for x in v)
    if isinstance(v, dict):
        return "; ".join(f"{k}: {x}" for k, x in v.items())
    return v


def build_excel_bytes(metrics_df: pd.DataFrame, winner_names: list,
                      analyses: dict, insights: dict) -> bytes:
    """Build a multi-sheet Excel: performance, winner breakdowns, and strategy."""
    import openpyxl
    from openpyxl.styles import Alignment, Font, PatternFill

    header_fill = PatternFill("solid", fgColor="2F75B6")
    header_font = Font(bold=True, color="FFFFFF")
    win_fill = PatternFill("solid", fgColor="D4EDDA")

    wb = openpyxl.Workbook()

    # Sheet 1: Performance
    ws = wb.active
    ws.title = "Performance"
    cols = ["ad_name", "spend", "impressions", "clicks", "conversions", "revenue",
            "CTR", "CVR", "CPA", "ROAS", "AOV"]
    ws.append([c.upper() if len(c) <= 4 else c.replace("_", " ").title() for c in cols])
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
    for _, row in metrics_df.iterrows():
        ws.append([_cell(row.get(c, "")) for c in cols])
        if row["ad_name"] in winner_names:
            for cell in ws[ws.max_row]:
                cell.fill = win_fill
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    for col in ws.columns:
        width = max((len(str(c.value or "")) for c in col), default=0)
        ws.column_dimensions[col[0].column_letter].width = min(width + 3, 40)

    # Sheet 2: Winner Analysis
    ws2 = wb.create_sheet("Winner Analysis")
    ws2.append(["Ad", "Hook Angle", "Why It Grabs Them", "Main Message",
                "Must-Have Positioning", "Objections Handled",
                "What Video Shows", "What Script Says", "Triggers"])
    for cell in ws2[1]:
        cell.font = header_font
        cell.fill = header_fill
    for name in winner_names:
        a = analyses.get(name, {})
        hook_style = a.get("hook_style", {})
        core_angle = a.get("core_angle_mechanism", {})
        visual_script = a.get("visual_script_structure", {})
        ws2.append([_cell(v) for v in (
            name, hook_style.get("angle", ""), hook_style.get("why_it_grabs_them", ""),
            core_angle.get("main_message", ""), core_angle.get("must_have_positioning", ""),
            a.get("objections_handled", []),
            visual_script.get("what_video_shows", ""),
            visual_script.get("what_transcript_says", ""),
            a.get("psychological_triggers", []),
        )])
    for col in ws2.columns:
        width = max((len(str(c.value or "")) for c in col), default=0)
        ws2.column_dimensions[col[0].column_letter].width = min(width + 3, 50)

    # Sheet 3: Strategy (patterns + new concepts)
    ws3 = wb.create_sheet("Strategy")
    ws3.append(["New Ad Concept", "Hook", "Angle", "Format", "Script Outline",
                "Why It Works", "What To Test"])
    for cell in ws3[1]:
        cell.font = header_font
        cell.fill = header_fill
    for c in insights.get("new_ad_concepts", []):
        ws3.append([_cell(v) for v in (
            c.get("concept_name", ""), c.get("hook", ""), c.get("angle", ""),
            c.get("format", ""), c.get("script_outline", ""),
            c.get("why_it_should_work", ""), c.get("what_to_test", ""),
        )])
    for col in ws3.columns:
        width = max((len(str(c.value or "")) for c in col), default=0)
        ws3.column_dimensions[col[0].column_letter].width = min(width + 3, 50)

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


# --------------------------------------------------------------------------- #
# Markdown report (Google-Doc-style)
# --------------------------------------------------------------------------- #
def build_markdown_report(metrics_df: pd.DataFrame, winner_names: list,
                          analyses: dict, insights: dict, metric: str) -> str:
    """Produce a shareable written report of the analysis."""
    lines = ["# Ad Creative Performance Report", ""]
    lines.append(f"Winners selected by **{metric}**. "
                 f"{len(winner_names)} winning ads analyzed out of {len(metrics_df)} total.")
    lines.append("")

    lines.append("## Winning Ads")
    for name in winner_names:
        row = metrics_df[metrics_df["ad_name"] == name].iloc[0]
        a = analyses.get(name, {})
        hook_style = a.get("hook_style", {})
        core_angle = a.get("core_angle_mechanism", {})
        visual_script = a.get("visual_script_structure", {})
        lines.append(f"### {name}")
        lines.append(f"- ROAS {row['ROAS']} · CTR {row['CTR']}% · "
                     f"CPA {row['CPA']} · Spend {row['spend']:,.0f}")
        lines.append(f"- **Hook angle:** {_cell(hook_style.get('angle',''))}")
        lines.append(f"- **Why it grabs them:** {_cell(hook_style.get('why_it_grabs_them',''))}")
        lines.append(f"- **Main message:** {_cell(core_angle.get('main_message',''))}")
        lines.append(f"- **Must-have positioning:** {_cell(core_angle.get('must_have_positioning',''))}")
        if a.get("objections_handled"):
            lines.append(f"- **Objections handled:** {_cell(a['objections_handled'])}")
        lines.append(f"- **What video shows:** {_cell(visual_script.get('what_video_shows',''))}")
        lines.append(f"- **What script says:** {_cell(visual_script.get('what_transcript_says',''))}")
        if visual_script.get("pacing_notes"):
            lines.append(f"- **Pacing notes:** {_cell(visual_script['pacing_notes'])}")
        if a.get("psychological_triggers"):
            lines.append(f"- **Triggers:** {_cell(a['psychological_triggers'])}")
        lines.append("")

    wp = insights.get("winning_patterns", {})
    lines.append("## What's Working (Patterns)")
    for label, key in [("Hooks", "common_hooks"), ("Angles", "common_angles"),
                       ("Formats", "common_formats"), ("Triggers", "common_triggers")]:
        vals = wp.get(key, [])
        if vals:
            lines.append(f"- **{label}:** {_cell(vals)}")
    if insights.get("why_it_works"):
        lines.append("")
        lines.append(f"**Why it works:** {_cell(insights['why_it_works'])}")
    lines.append("")

    ai = insights.get("audience_insights", {})
    lines.append("## Audience Insights")
    for label, key in [("Status drivers", "status_drivers"),
                       ("Prestige desires", "prestige_desires"),
                       ("Key objections", "key_objections")]:
        vals = ai.get(key, [])
        if vals:
            lines.append(f"- **{label}:** {_cell(vals)}")
    lines.append("")

    lines.append("## New Creative Concepts to Test")
    for i, c in enumerate(insights.get("new_ad_concepts", []), 1):
        lines.append(f"### {i}. {_cell(c.get('concept_name',''))}")
        lines.append(f"- **Hook:** {_cell(c.get('hook',''))}")
        lines.append(f"- **Angle:** {_cell(c.get('angle',''))} · **Format:** {_cell(c.get('format',''))}")
        lines.append(f"- **Script outline:** {_cell(c.get('script_outline',''))}")
        lines.append(f"- **Why it should work:** {_cell(c.get('why_it_should_work',''))}")
        lines.append(f"- **What to test:** {_cell(c.get('what_to_test',''))}")
        lines.append("")

    recs = insights.get("testing_recommendations", [])
    if recs:
        lines.append("## Testing Recommendations")
        for r in recs:
            lines.append(f"- {_cell(r)}")
        lines.append("")

    return "\n".join(lines)
